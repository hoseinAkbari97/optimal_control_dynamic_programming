import numpy as np
from collections import defaultdict
import openpyxl

# -----------------------------
# Discretization of state/control
# -----------------------------

def discretize_spaces(x_step, u_step):
    x_min, x_max = 0.0, 1.5
    u_min, u_max = -1.0, 1.0

    x_vals = np.arange(x_min, x_max + 1e-9, x_step)
    u_vals = np.arange(u_min, u_max + 1e-9, u_step)
    return x_vals, u_vals

# -----------------------------
# Time grid (k = 0, 1, ..., N)
# -----------------------------

def make_time_grid(N, T=2.0):
    k_grid = np.arange(N + 1)
    dt = T / N if N > 0 else T
    return k_grid, dt

# -----------------------------
# Problem-specific ingredients (given by user)
# -----------------------------

def build_x_index(x_vals, tol=1e-9):
    return {round(float(v), 10): i for i, v in enumerate(x_vals)}


def valid_next_index(x, u, x_index_map, x_min=0.0, x_max=1.5):
    x_next = round(float(x + u), 10)
    if x_next < x_min - 1e-12 or x_next > x_max + 1e-12:
        return None, None
    i_next = x_index_map.get(x_next, None)
    return x_next, i_next

# Costs

def stage_cost(u_k):
    return 2.0 * (u_k ** 2)


def terminal_step_cost(x_next, u_last):
    return (x_next ** 2) + 2.0 * (u_last ** 2)

# -----------------------------
# DP Solver
# -----------------------------

def solve_dp(x_step=0.5, u_step=0.5, N=2, T=2.0):
    x_vals, u_vals = discretize_spaces(x_step, u_step)
    k_grid, dt = make_time_grid(N, T)
    nx, nu = len(x_vals), len(u_vals)

    x_index_map = build_x_index(x_vals)

    J = np.full((N, nx), np.inf)
    pi = np.full((N, nx), -1, dtype=int)

    step_details = [defaultdict(list) for _ in range(N)]

    # Last step
    if N >= 1:
        k = N - 1
        for i, x in enumerate(x_vals):
            best_cost = np.inf
            best_u_idx = -1
            for uj, u in enumerate(u_vals):
                x_next, i_next = valid_next_index(x, u, x_index_map)
                if i_next is None:
                    continue
                total = terminal_step_cost(x_next, u)
                step_details[k][i].append({
                    'u': float(u),
                    'x_next': float(x_next),
                    'immediate_cost': float((x_next ** 2) + 2.0 * (u ** 2)),
                    'next_cost': 0.0,
                    'total_cost': float(total),
                })
                if total < best_cost - 1e-12:
                    best_cost = total
                    best_u_idx = uj
            if best_u_idx != -1:
                J[k, i] = best_cost
                pi[k, i] = best_u_idx

    # Previous steps
    for k in range(N - 2, -1, -1):
        for i, x in enumerate(x_vals):
            best_cost = np.inf
            best_u_idx = -1
            for uj, u in enumerate(u_vals):
                x_next, i_next = valid_next_index(x, u, x_index_map)
                if i_next is None:
                    continue
                immediate = stage_cost(u)
                next_c = J[k + 1, i_next]
                total = immediate + next_c
                step_details[k][i].append({
                    'u': float(u),
                    'x_next': float(x_next),
                    'immediate_cost': float(immediate),
                    'next_cost': float(next_c),
                    'total_cost': float(total),
                })
                if total < best_cost - 1e-12:
                    best_cost = total
                    best_u_idx = uj
            if best_u_idx != -1:
                J[k, i] = best_cost
                pi[k, i] = best_u_idx

    return {
        'x_vals': x_vals,
        'u_vals': u_vals,
        'N': N,
        'T': T,
        'dt': dt,
        'J': J,
        'pi': pi,
        'details': step_details,
    }

# -----------------------------
# Export to Excel
# -----------------------------

def export_to_excel(result, filename="dp_solution.xlsx"):
    wb = openpyxl.Workbook()
    x_vals = result['x_vals']
    u_vals = result['u_vals']
    J = result['J']
    pi = result['pi']
    details = result['details']
    N = result['N']

    for k in range(N - 1, -1, -1):
        ws = wb.create_sheet(title=f"Step_{k}")
        ws.append(["x", "u", "x_next", "immediate_cost", "next_cost", "total_cost", "j*", "u*"])
        for i, x in enumerate(x_vals):
            opts = details[k].get(i, [])
            jstar = J[k, i]
            u_idx = pi[k, i]
            ustar = u_vals[u_idx] if u_idx >= 0 else None
            if not opts:
                ws.append([x, None, None, None, None, None, jstar, ustar])
            else:
                for o in opts:
                    ws.append([
                        x,
                        o['u'],
                        o['x_next'],
                        o['immediate_cost'],
                        o['next_cost'],
                        o['total_cost'],
                        jstar,
                        ustar
                    ])

    # remove default sheet
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    wb.save(filename)
    print(f"DP solution exported to {filename}")

# -----------------------------
# CLI demo
# -----------------------------
if __name__ == "__main__":
    x_step = float(input("Enter step size for x (e.g. 0.5): "))
    u_step = float(input("Enter step size for u (e.g. 0.5): "))
    N = int(input("Enter number of steps N (e.g. 4): "))

    result = solve_dp(x_step=x_step, u_step=u_step, N=N, T=2.0)
    export_to_excel(result, filename="dp_solution.xlsx")
